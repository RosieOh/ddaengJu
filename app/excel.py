# -*- coding: utf-8 -*-
"""엑셀 리포트 생성.

이미지가 아니라 **엑셀 네이티브 차트**를 심는다. 받는 사람이 엑셀에서
그대로 축·기간을 바꿔 볼 수 있고, 파일 하나로 공유된다.

시트 구성
  요약       KPI + 키워드 × 업체 현재 순위 (색조 조건부서식) + 직전 대비 변화
  순위추이   키워드별 라인차트 (Y축 역방향 — 1위가 위)
  시간대별   시간대 평균 순위 막대차트
  변동성     순위가 불안정한 조합
  순위이력   원본 데이터
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# viz.css 의 순차 램프와 같은 계열(파랑). 상위=진함.
RAMP_DARK = "FF184F95"
RAMP_MID = "FF6DA7EC"
RAMP_LIGHT = "FFCDE2FB"

HEAD_FILL = PatternFill("solid", fgColor="FFF2F2F2")
HEAD_FONT = Font(bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13)
MUTED_FONT = Font(size=9, color="FF898781")
GOOD_FONT = Font(size=10, color="FF006300", bold=True)
BAD_FONT = Font(size=10, color="FFD03B3B", bold=True)
THIN = Side(style="thin", color="FFE1E0D9")
BORDER = Border(bottom=THIN)

OUT_LABEL = "순위밖"


def _style_header(ws, row: int, ncols: int, start: int = 1) -> None:
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")


def _widths(ws, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _rank_cell(v):
    return OUT_LABEL if v is None else v


def build_workbook(analysis: dict, raw_rows: Sequence[dict], window_hours: int) -> io.BytesIO:
    wb = Workbook()
    targets: list[str] = analysis["targets"]
    primary: str = analysis["primary"]
    first_page: int = analysis["first_page"]

    _sheet_summary(wb.active, analysis, targets, primary, first_page, window_hours)
    _sheet_trend(wb.create_sheet("순위추이"), analysis, targets)
    _sheet_hourly(wb.create_sheet("시간대별"), analysis, targets, primary)
    _sheet_volatility(wb.create_sheet("변동성"), analysis)
    _sheet_raw(wb.create_sheet("순위이력"), raw_rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ------------------------------------------------------------------- 요약


def _sheet_summary(ws, analysis, targets, primary, first_page, window_hours) -> None:
    ws.title = "요약"
    kpi = analysis.get("kpi") or {}

    ws["A1"] = "네이버 파워링크 노출순위 리포트"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        f"기준 업체 {primary} · 최근 {window_hours}시간 · "
        f"생성 {datetime.now():%Y-%m-%d %H:%M} · 모바일 파워링크 기준(1페이지 {first_page}개)"
    )
    ws["A2"].font = MUTED_FONT

    row = 4
    ws.cell(row=row, column=1, value="핵심 지표").font = Font(bold=True, size=11)
    row += 1
    # 증감은 직전 기록이 있는 키워드끼리만 비교한다(analytics.kpi 와 같은 기준).
    compared = kpi.get("compared", 0)

    def pair(key: str):
        return (kpi.get(f"{key}_now", 0) - kpi.get(f"{key}_prev", 0)) if compared else None

    cards = [
        ("평균 순위", "—" if kpi.get("avg_rank") is None else kpi["avg_rank"],
         kpi.get("avg_rank_delta") if compared else None, "위", False),
        (f"첫 페이지({first_page}위 내) 노출", kpi.get("first_page"), pair("first_page"), "개", False),
        ("상위 3위 이내", kpi.get("top3"), pair("top3"), "개", False),
        ("순위밖(미노출)", kpi.get("out"), pair("out"), "개", True),
    ]
    ws.cell(row=row, column=1, value="지표")
    ws.cell(row=row, column=2, value="현재")
    ws.cell(row=row, column=3, value="직전 대비")
    _style_header(ws, row, 3)
    for label, value, delta, unit, invert in cards:
        row += 1
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        cell = ws.cell(row=row, column=3)
        if delta is None:
            cell.value = "비교할 직전 기록 없음"
            cell.font = MUTED_FONT
        elif delta == 0:
            cell.value = "변화 없음"
            cell.font = MUTED_FONT
        else:
            good = (delta < 0) if invert else (delta > 0)
            cell.value = f"{'▲' if delta > 0 else '▼'} {abs(delta)}{unit}"
            cell.font = GOOD_FONT if good else BAD_FONT

    # ── 키워드 × 업체 현재 순위
    row += 3
    head_row = row
    ws.cell(row=row, column=1, value="키워드")
    for i, t in enumerate(targets):
        ws.cell(row=row, column=2 + i, value=t)
    ws.cell(row=row, column=2 + len(targets), value="전체 광고수")
    ws.cell(row=row, column=3 + len(targets), value=f"{primary} 직전 대비")
    _style_header(ws, row, len(targets) + 3)

    for entry in analysis["current"]:
        row += 1
        ws.cell(row=row, column=1, value=entry["keyword"])
        for i, t in enumerate(targets):
            info = entry["ranks"].get(t) or {}
            ws.cell(row=row, column=2 + i, value=_rank_cell(info.get("rank")))
        ws.cell(row=row, column=2 + len(targets), value=entry.get("total_ads") or 0)
        delta = (entry["ranks"].get(primary) or {}).get("delta")
        cell = ws.cell(row=row, column=3 + len(targets))
        if delta is None:
            cell.value = "—"
            cell.font = MUTED_FONT
        elif delta == 0:
            cell.value = "변화 없음"
            cell.font = MUTED_FONT
        else:
            cell.value = f"{'▲' if delta > 0 else '▼'} {abs(delta)}"
            cell.font = GOOD_FONT if delta > 0 else BAD_FONT

    # 순위는 작을수록 좋으므로 색조를 뒤집어(작은 값=진한 파랑) 적용한다.
    if row > head_row:
        rng = f"B{head_row + 1}:{get_column_letter(1 + len(targets))}{row}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min", start_color=RAMP_DARK,
                mid_type="percentile", mid_value=50, mid_color=RAMP_MID,
                end_type="max", end_color=RAMP_LIGHT,
            ),
        )

    row += 2
    ws.cell(row=row, column=1, value="※ 순위는 3회 측정 후 중앙값입니다. 파워링크는 호출마다 광고가 로테이션되어 순위가 흔들립니다.").font = MUTED_FONT

    _widths(ws, {1: 22, **{2 + i: 16 for i in range(len(targets) + 2)}})
    ws.freeze_panes = ws.cell(row=head_row + 1, column=2)


# ---------------------------------------------------------------- 순위추이
def _sheet_trend(ws, analysis, targets) -> None:
    """키워드별 데이터 블록 + 라인차트. Y축은 1위가 위로 오도록 뒤집는다."""
    ws["A1"] = "키워드별 순위 추이"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Y축은 위쪽이 상위 노출입니다. 순위밖은 빈칸으로 두어 선이 끊어집니다."
    ws["A2"].font = MUTED_FONT

    row = 4
    trend: dict = analysis.get("trend") or {}
    for keyword, per_target in trend.items():
        stamps = sorted({p[0] for series in per_target.values() for p in series})
        if len(stamps) < 2:
            continue

        ws.cell(row=row, column=1, value=keyword).font = Font(bold=True, size=11)
        row += 1
        head = row
        ws.cell(row=row, column=1, value="조회시각")
        cols = [t for t in targets if per_target.get(t)]
        for i, t in enumerate(cols):
            ws.cell(row=row, column=2 + i, value=t)
        _style_header(ws, row, len(cols) + 1)

        lookup = {t: {p[0]: p[1] for p in per_target.get(t, [])} for t in cols}
        for stamp in stamps:
            row += 1
            ws.cell(row=row, column=1, value=datetime.fromisoformat(stamp).strftime("%m-%d %H:%M"))
            for i, t in enumerate(cols):
                v = lookup[t].get(stamp)
                if v is not None:
                    ws.cell(row=row, column=2 + i, value=v)

        chart = LineChart()
        chart.title = f"{keyword} 순위 추이"
        chart.style = 2
        chart.height = 7.5
        chart.width = 20
        chart.y_axis.title = "노출순위"
        chart.y_axis.scaling.orientation = "maxMin"   # 1위가 위
        chart.x_axis.title = "조회시각"
        data = Reference(ws, min_col=2, max_col=1 + len(cols), min_row=head, max_row=row)
        cats = Reference(ws, min_col=1, min_row=head + 1, max_row=row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        for s in chart.series:
            s.smooth = False
        ws.add_chart(chart, f"{get_column_letter(len(cols) + 3)}{head}")

        row += max(3, 16 - (row - head))

    if row <= 4:
        ws.cell(row=4, column=1, value="추이를 그리려면 같은 키워드로 2회 이상 수집되어야 합니다.").font = MUTED_FONT

    _widths(ws, {1: 16, 2: 14, 3: 14, 4: 14, 5: 14, 6: 14})


# ---------------------------------------------------------------- 시간대별
def _sheet_hourly(ws, analysis, targets, primary) -> None:
    ws["A1"] = "시간대별 평균 순위"
    ws["A1"].font = TITLE_FONT
    hourly = analysis.get("hourly") or {}
    best, worst = hourly.get("best_hour"), hourly.get("worst_hour")
    note = "자동 수집이 여러 시간대에 걸쳐 돌아야 의미 있는 값이 나옵니다."
    if best and worst:
        note = (f"{primary} 기준 — 가장 좋은 시간대 {best['hour']:02d}시(평균 {best['avg']}위) · "
                f"가장 밀리는 시간대 {worst['hour']:02d}시(평균 {worst['avg']}위)")
    ws["A2"] = note
    ws["A2"].font = MUTED_FONT

    head = 4
    ws.cell(row=head, column=1, value="시간")
    for i, t in enumerate(targets):
        ws.cell(row=head, column=2 + i, value=t)
    ws.cell(row=head, column=2 + len(targets), value="표본수")
    _style_header(ws, head, len(targets) + 2)

    rows_written = 0
    row = head
    for h in hourly.get("hours", []):
        vals = [hourly["series"][t][h] for t in targets]
        if all(v is None for v in vals):
            continue
        row += 1
        rows_written += 1
        ws.cell(row=row, column=1, value=f"{h:02d}시")
        for i, v in enumerate(vals):
            if v is not None:
                ws.cell(row=row, column=2 + i, value=v)
        ws.cell(row=row, column=2 + len(targets), value=hourly["counts"][h])

    if rows_written >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = "시간대별 평균 노출순위 (낮을수록 상위)"
        chart.height = 8
        chart.width = 22
        chart.y_axis.title = "평균 순위"
        chart.y_axis.scaling.orientation = "maxMin"
        chart.x_axis.title = "시간"
        chart.gapWidth = 40
        data = Reference(ws, min_col=2, max_col=1 + len(targets), min_row=head, max_row=row)
        cats = Reference(ws, min_col=1, min_row=head + 1, max_row=row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"{get_column_letter(len(targets) + 4)}{head}")

    _widths(ws, {1: 10, **{2 + i: 16 for i in range(len(targets) + 1)}})


# ----------------------------------------------------------------- 변동성


def _sheet_volatility(ws, analysis) -> None:
    ws["A1"] = "순위 변동성"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "표준편차가 클수록 순위가 자주 출렁입니다. 입찰가 점검이 필요한 조합입니다."
    ws["A2"].font = MUTED_FONT

    head = 4
    cols = ["키워드", "업체", "표준편차", "최고순위", "최저순위", "표본수", "순위밖 횟수"]
    for i, c in enumerate(cols):
        ws.cell(row=head, column=1 + i, value=c)
    _style_header(ws, head, len(cols))

    row = head
    for v in analysis.get("volatility", []):
        row += 1
        for i, key in enumerate(["keyword", "target", "stdev", "min", "max", "samples", "out_count"]):
            ws.cell(row=row, column=1 + i, value=v[key])

    if row == head:
        ws.cell(row=head + 1, column=1, value="표본이 3회 미만이라 아직 계산할 수 없습니다.").font = MUTED_FONT

    _widths(ws, {1: 20, 2: 18, 3: 12, 4: 12, 5: 12, 6: 10, 7: 14})


# ---------------------------------------------------------------- 순위이력


def _sheet_raw(ws, rows) -> None:
    ws["A1"] = "순위 이력 (원본)"
    ws["A1"].font = TITLE_FONT

    head = 3
    cols = ["조회시각", "키워드", "업체", "순위", "전체광고수", "순위변동", "측정값", "수집방식"]
    for i, c in enumerate(cols):
        ws.cell(row=head, column=1 + i, value=c)
    _style_header(ws, head, len(cols))

    row = head
    for r in rows:
        row += 1
        ws.cell(row=row, column=1, value=r["collected_at"].replace("T", " "))
        ws.cell(row=row, column=2, value=r["keyword"])
        ws.cell(row=row, column=3, value=r["target"])
        ws.cell(row=row, column=4, value=_rank_cell(r["rank"]))
        ws.cell(row=row, column=5, value=r["total_ads"])
        ws.cell(row=row, column=6, value="변동" if r["unstable"] else "")
        ws.cell(row=row, column=7, value=r["samples"])
        ws.cell(row=row, column=8, value="자동" if r["source"] == "schedule" else "수동")

    _widths(ws, {1: 20, 2: 20, 3: 18, 4: 10, 5: 12, 6: 10, 7: 14, 8: 10})
    ws.freeze_panes = ws.cell(row=head + 1, column=1)
